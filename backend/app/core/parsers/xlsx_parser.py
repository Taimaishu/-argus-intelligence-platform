"""XLSX spreadsheet parser using openpyxl."""

from openpyxl import load_workbook
from .base import BaseParser, ParsedDocument


class XLSXParser(BaseParser):
    """Parser for XLSX spreadsheets."""

    def supports_file_type(self, file_extension: str) -> bool:
        """Check if file extension is .xlsx"""
        return file_extension.lower() in [".xlsx", ".xls"]

    def parse(self, file_path: str) -> ParsedDocument:
        """
        Parse XLSX spreadsheet and extract data.

        Args:
            file_path: Path to XLSX file

        Returns:
            ParsedDocument with extracted content
        """
        try:
            workbook = load_workbook(file_path, data_only=True)

            # Extract metadata
            props = workbook.properties
            title = props.title or "Spreadsheet"
            author = props.creator

            full_text = ""
            sections = []

            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                sheet_text = f"Sheet: {sheet_name}\n\n"

                # Get all rows with values
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    # Skip empty rows
                    if any(cell is not None for cell in row):
                        row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                        rows.append(row_text)
                        sheet_text += row_text + "\n"

                if rows:
                    sections.append({
                        "heading": sheet_name,
                        "content": sheet_text
                    })

                    full_text += sheet_text + "\n\n"

            workbook.close()

            return ParsedDocument(
                text=self.clean_text(full_text),
                title=title,
                author=author,
                metadata={
                    "sheet_count": len(workbook.sheetnames),
                    "sheet_names": workbook.sheetnames,
                    "created": str(props.created) if props.created else None,
                    "modified": str(props.modified) if props.modified else None,
                },
                sections=sections
            )

        except Exception as e:
            raise Exception(f"Failed to parse XLSX: {str(e)}")
